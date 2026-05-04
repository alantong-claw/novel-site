from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import sys

ROOT = Path('/home/alantong/ai-work/excel')
HELPER_DIR = Path('/home/alantong/ai-work/scripts')
if str(HELPER_DIR) not in sys.path:
    sys.path.append(str(HELPER_DIR))
from office_output_path import output_path
OUT = output_path('2026-03-22-edge-ai-tv-openclaw-report', '2026-03-22-edge-ai-tv-openclaw-report-v1.xlsx', 'OUTPUT_XLSX')

wb = Workbook()
ws = wb.active
ws.title = 'Summary'

rows = [
    ['主題', '15 TOPS Edge AI + TV 控制 + OpenClaw 應用討論'],
    ['一句話結論', '應做有邊界的 edge AV / control appliance，不要做萬用 AI 電視大腦。'],
    ['市場價值', '本地控制、低延遲、隱私、可離線、硬體整合'],
    ['最有潛力場景', '會議室/教室、數位看板/kiosk、安防顯示牆、Prosumer AV、無障礙控制'],
    ['技術主張', '最佳是 edge box 自己就是播放來源；折衷是控制 TV 但不吃進所有影音'],
    ['不建議主路徑', '任意 HDMI capture，容易踩 EDID / latency / HDCP 地雷'],
    ['OpenClaw 安裝', 'Linux 直接安裝最推薦：64-bit Linux + Node 22+/24 + onboard --install-daemon'],
    ['主要限制', 'DRM/HDCP、CEC 相容性、客廳 support 成本、AI oversell'],
    ['MVP 建議', 'Linux mini-PC / SBC + HDMI output + USB HDMI-CEC + mic/speaker'],
    ['Go / No-Go', 'Go: 環境可控、矩陣可驗證；No-Go: 想通吃所有 TV / HDMI / 受保護內容'],
]
for row in rows:
    ws.append(row)

header_fill = PatternFill('solid', fgColor='1F4E78')
header_font = Font(color='FFFFFF', bold=True)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
    for cell in row:
        cell.alignment = Alignment(vertical='top', wrap_text=True)

ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 88

ws2 = wb.create_sheet('UseCases')
ws2.append(['類別', '可做的事', '價值'])
for row in [
    ['會議室 / 教室', '切來源、開關顯示器、啟停錄影/串流', '降低操作摩擦'],
    ['數位看板 / kiosk', '來源切換、fallback content、本地分析', '本地可靠、低延遲'],
    ['安防 / 營運顯示牆', '事件驅動顯示切換、本地 triage', '反應更快'],
    ['Prosumer AV', 'scene/source automation、語音工作流', '減少手動 glue code'],
    ['無障礙控制', '語音控制 TV/display/source/volume', '提高可用性'],
]:
    ws2.append(row)
for cell in ws2[1]:
    cell.fill = header_fill
    cell.font = header_font
for col, width in {'A':20, 'B':40, 'C':28}.items():
    ws2.column_dimensions[col].width = width

ws3 = wb.create_sheet('Recommendations')
ws3.append(['項目', '建議'])
for row in [
    ['產品定位', '專用 edge AV / control appliance'],
    ['控制主路徑', 'HDMI-CEC + fallback control'],
    ['部署模式', 'edge-first 做本地 I/O，central assist 做管理'],
    ['MVP 優先', '先做 local playback / local UI，不碰 arbitrary HDMI ingest'],
    ['商業邊界', '避免 unmanaged consumer 市場起手'],
]:
    ws3.append(row)
for cell in ws3[1]:
    cell.fill = header_fill
    cell.font = header_font
ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 80

wb.save(OUT)
print(OUT)
