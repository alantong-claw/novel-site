from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import sys

ROOT = Path('/home/alantong/ai-work/excel')
HELPER_DIR = Path('/home/alantong/ai-work/scripts')
if str(HELPER_DIR) not in sys.path:
    sys.path.append(str(HELPER_DIR))
from office_output_path import output_path
OUT = output_path('2026-03-22-hsinchu-science-vs-gifted', '2026-03-22-hsinchu-science-vs-gifted-v1.xlsx', 'OUTPUT_XLSX')

wb = Workbook()
ws = wb.active
ws.title = 'Compare'
header_fill = PatternFill('solid', fgColor='1F4E78')
header_font = Font(color='FFFFFF', bold=True)

rows = [
    ['面向', '科學班', '資優班'],
    ['核心定位', 'STEM 集中加速', '高能力但保留彈性'],
    ['優勢', '數理更集中、同儕更偏 STEM、研究/競賽資源更近', '彈性高、全面發展、保留更多選擇'],
    ['風險', '壓力集中、容易 burnout、可能太早鎖定', '若已明確偏 STEM 可能覺得不夠集中，較依賴自律'],
    ['適合誰', '已明確偏 STEM、喜歡深挖與高密度訓練者', '很強但方向未定、重視彈性與整體發展者'],
    ['一句話', '確定走 STEM，就讓資源更集中', '還在探索，就先保留選擇權'],
]
for row in rows:
    ws.append(row)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
    for cell in row:
        cell.alignment = Alignment(vertical='top', wrap_text=True)
for col, width in {'A':16, 'B':46, 'C':46}.items():
    ws.column_dimensions[col].width = width

ws2 = wb.create_sheet('Recommendation')
ws2.append(['情況', '建議'])
for row in [
    ['已非常明確偏 STEM', '傾向科學班'],
    ['很強但方向未定', '傾向資優班'],
    ['想走研究/競賽/醫工理科', '科學班較有利'],
    ['想保留跨域與整體發展', '資優班較穩健'],
    ['真正關鍵', '不是名聲，而是適配度'],
]:
    ws2.append(row)
for cell in ws2[1]:
    cell.fill = header_fill
    cell.font = header_font
ws2.column_dimensions['A'].width = 28
ws2.column_dimensions['B'].width = 54

wb.save(OUT)
print(OUT)
