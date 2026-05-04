from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import sys

ROOT = Path('/home/alantong/ai-work/excel')
HELPER_DIR = Path('/home/alantong/ai-work/scripts')
if str(HELPER_DIR) not in sys.path:
    sys.path.append(str(HELPER_DIR))
from office_output_path import output_path
OUT = output_path('2026-03-23-tv-edge-15tops-market', '2026-03-23-tv-edge-15tops-market-v1.xlsx', 'OUTPUT_XLSX')

wb = Workbook()
ws = wb.active
ws.title = 'CrossAspect'
header_fill = PatternFill('solid', fgColor='1F4E78')
header_font = Font(color='FFFFFF', bold=True)
rows = [
    ['面向', '正方收斂', '反方收斂'],
    ['產品 / 應用', 'TV experience processor；先打 dialogue / volume / captions', '需要單一高頻 wedge；不能只靠 tech demo'],
    ['GTM / 商業推廣', '先切 operator-managed TV/STB、hospitality、managed display；賣 solution stack', 'buyer/ROI/channel 若不清，產品會商業無家可歸'],
    ['競爭 / 風險', '補在 TV SoC 與 cloud AI 間的 gap；走 companion / controlled pipeline 路線', 'integrated SoC、cloud、support burden、content rights 都可能吃掉價值'],
    ['最終建議', '先做窄場景、窄 buyer、窄 use case 的 edge media AI platform', '若拿掉 AI 標籤後價值不成立，就不該大推'],
]
for row in rows:
    ws.append(row)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
    for cell in row:
        cell.alignment = Alignment(vertical='top', wrap_text=True)
for col, width in {'A':18, 'B':50, 'C':50}.items():
    ws.column_dimensions[col].width = width

ws2 = wb.create_sheet('Recommendation')
for row in [
    ['項目', '建議'],
    ['最佳第一波產品主張', 'TV experience processor / edge media AI platform'],
    ['最佳第一波功能', 'dialogue enhancement、volume leveling、smarter subtitles'],
    ['最佳第一波市場', 'operator-managed TV/STB/IPTV、hospitality、managed displays'],
    ['最佳賣法', 'reference platform + BSP + SDK + models + OTA/support'],
    ['不要做', 'generic AI TV chip、mass-market TV first、過度擬人化 AI 敘事'],
    ['最重要驗證', '單一高頻 use case、ROI、integration feasibility、design partners'],
]:
    ws2.append(row)
for cell in ws2[1]:
    cell.fill = header_fill
    cell.font = header_font
ws2.column_dimensions['A'].width = 24
ws2.column_dimensions['B'].width = 78
for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=2):
    for cell in row:
        cell.alignment = Alignment(vertical='top', wrap_text=True)

wb.save(OUT)
print(OUT)
