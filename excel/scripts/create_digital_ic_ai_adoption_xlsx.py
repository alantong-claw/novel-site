from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import sys

ROOT = Path('/home/alantong/ai-work/excel')
HELPER_DIR = Path('/home/alantong/ai-work/scripts')
if str(HELPER_DIR) not in sys.path:
    sys.path.append(str(HELPER_DIR))
from office_output_path import output_path
OUT = output_path('2026-03-23-digital-ic-ai-adoption', '2026-03-23-digital-ic-ai-adoption-v1.xlsx', 'OUTPUT_XLSX')

wb = Workbook()
ws = wb.active
ws.title = 'ProCon'
header_fill = PatternFill('solid', fgColor='1F4E78')
header_font = Font(color='FFFFFF', bold=True)
rows = [
    ['面向', '支持推廣（A）', '反方質疑（B）'],
    ['核心立場', '先從低風險、高摩擦、可審查任務切入', '即使縮範圍，也未必證明 net value'],
    ['優先場景', 'log triage、spec search、scripts、docs、review prep、testbench scaffolding', '這些也可能因 automation bias 與錯誤摘要造成問題'],
    ['高風險場景', 'critical RTL / STA / CDC waiver / ECO / signoff 不應早期導入', '若邊界不清或無法技術性強制， rollout 不成熟'],
    ['價值主張', '減少低價值重複工作、加快 onboarding 與 triage', 'first draft 快不等於 trusted throughput 提升'],
    ['主要風險', 'hallucination、integration friction、security、ROI uncertainty', 'silent wrongness、context loss、review quality erosion、metrics gaming'],
    ['落地方法', 'risk-tier rollout + pilot + metrics + human review', 'pilot 容易被 cherry-pick、enthusiast bias、novelty effect 誤導'],
    ['關鍵治理', 'approved models, data classes, provenance, mandatory review', '若 audit trail、policy enforcement、negative-event logging 不足，就不應擴張'],
    ['Go 條件', '10–20% task-level 改善且品質不退步', '若 defect/rework/review burden 上升，應立刻暫停'],
    ['一句話', '先建立可信度，再談擴張', '先證明不傷品質與判斷力，再談採用'],
]
for row in rows:
    ws.append(row)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
    for cell in row:
        cell.alignment = Alignment(vertical='top', wrap_text=True)
for col, width in {'A':16, 'B':48, 'C':48}.items():
    ws.column_dimensions[col].width = width

ws2 = wb.create_sheet('Recommendation')
for row in [
    ['項目', '建議'],
    ['先推什麼', 'Regression/log triage、internal knowledge retrieval、scripts、docs、review support'],
    ['後推什麼', 'testbench/assertion scaffolding、RTL skeleton from reviewed templates'],
    ['先不要推', 'critical RTL、SDC/STA、CDC waiver、ECO、signoff judgment'],
    ['成功條件', '有可量測的 task-level ROI，且品質、安全、review burden 不惡化'],
    ['核心策略', '不是 AI evangelism，而是 controlled credibility play'],
]:
    ws2.append(row)
for cell in ws2[1]:
    cell.fill = header_fill
    cell.font = header_font
ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 72
for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=2):
    for cell in row:
        cell.alignment = Alignment(vertical='top', wrap_text=True)

wb.save(OUT)
print(OUT)
